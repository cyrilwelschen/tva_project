'__author__' == 'cyril'


import os
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

DATAPATH = "/home/cyril/Desktop/data_bix/"


def list_files(path=None):
    """
    List files and directories of path.
    :param path: path to directory for display
    :return list of directories and files sorted alphabetically.
    """
    path = analysed_path(path)
    file_list = os.listdir(path)
    file_list.sort()
    return file_list


def analysed_path(path):
    """
    Analyse if path is provided, and if so if its relative or not
    :param path: absolute or relative path or None
    :return usable absolute path
    """
    if not path:
        return DATAPATH
    elif path[:5] == "/home":
        return path
    else:
        return DATAPATH+path


def list_filetype(ending, path=None):
    """
    List files of specific type (i.e. ending) in directory.
    :param ending: string of ending with variable length. E.g. 'xls'
    :param path: absolute or relative path or None
    :return list of files with that endng.
    """
    all_files = list_files(path)
    xls = []
    for fl in all_files:
        if ending == fl[-len(ending):]:
            xls.append(fl)
    return xls


def list_xls(path=None):
    """
    List all excel files of old type.
    :param path: absolute, relative or None
    :return: list of xls files
    """
    return list_filetype("xls", path)


def list_xlsx(path=None):
    """
    List all excel files of new type.
    :param path: absolute, relative or None
    :return: list of xlsx files
    """
    return list_filetype("xlsx", path)


def list_all_excels(path=None, merge=True):
    """
    List(s) all excel files, old and new type.
    :param path: absolute, relative or None
    :param merge: whether to return old and new type excels in sparate list.
    Defaulte is returning a single list.
    :return: list or two lists of xls and xlsx files
    """
    if merge:
        return list_xls(path) + list_xlsx(path)
    else:
        return list_xls(path),  list_xlsx(path)


def convert_excels(path_to_xls):
    os.system("libreoffice --convert-to xlsx "+path_to_xls+" --headless")
    print("libreoffice --convert-to xlsx "+path_to_xls+" --headless")
    pass


def convert_all_excels(path):
    pass


def excels_by_regex(path):
    pass


def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1
    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
    return book1


def convert_xls_to_xlsx(path_to_file, keep_original=True):
    filename = path_to_file
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1
    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
    return book1
