'__author__' == 'cyril'


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from tva_project.utils import filesystem as fs


def datapath():
    """
    Info about default datapath
    :return string of default datapath
    """
    return fs.datapath()


def load_wb(filename):
    """
    Get list of worksheet names of an xlsx file.
    :param filename: path to file
    :return list of filenames
    """
    try:
        wb = load_workbook(filename, read_only=True)
    except FileNotFoundError:
        print("File {} not found.".format(filename))
    except:
        print("Faild to load file {}".format(filename))
        raise
    else:
        return wb


def get_sheetnames(workbook):
    """
    List of worksheets found for workbook.
    :param workbook: openpyxl workbook instance
    :return list of found worksheet names
    """
    return workbook.get_sheet_names()


def default_ws_names():
    """
    #TODO: add description
    """
    return ["ReZ-ReZType", "All_Data_ReZ"]


def find_important_ws(filename, overwrite_ws_name=None):
    """
    #todo: duplicate from below, rewrite the below
    """
    wb = load_wb(filename)
    name_proposal = overwrite_ws_name or default_ws_names()
    wb_sheet_names = get_sheetnames(wb)
    name = None
    if isinstance(name_proposal, str):
        if name_proposal in wb_sheet_names:
            name = name_proposal
    else:
        for name_prop in name_proposal:
            if name_prop in wb_sheet_names:
                name = name_prop
    try:
        ws = wb[name]
    except NameError:
        print("Couldn't find worksheet {}. Found {}".format(name,
                                                            get_sheetnames(wb)))
    else:
        return ws


def try_find_column_names(filename, overwrite_ws_name=None):
    """
    Try to find relevant worksheet colunm names.
    :param filename: path to filename.
    :param overwrite_ws_name: Optional name of worksheet to use instead of
    default list: default_ws_names()
    :return list of proposed column names.
    """
    wb = load_wb(filename)
    name_proposal = overwrite_ws_name or default_ws_names()
    wb_sheet_names = get_sheetnames(wb)
    name = None
    if isinstance(name_proposal, str):
        if name_proposal in wb_sheet_names:
            name = name_proposal
    else:
        for name_prop in name_proposal:
            if name_prop in wb_sheet_names:
                name = name_prop
    try:
        ws = wb[name]
    except NameError:
        print("Couldn't find worksheet {}. Found {}".format(name,
                                                            get_sheetnames(wb)))
    else:
        return [c.value for c in ws[1]]


# todo: slim list vs string
def find_columns_by_regex(filename, regex):
    """
    Find available column names from pattern.
    :param filename: file to look at
    :param regex: substring (pattern) to look for
    :return List of tuples (columnnames, index) where columnames contains regex.
    """
    all_column_names = try_find_column_names(filename)
    findings = []
    for i, n in enumerate(all_column_names):
        if isinstance(regex, str):
            if n is not None:
                if regex in n:
                    findings.append((get_column_letter(i+1), n))
        elif isinstance(regex, list):
            found = True
            for r in regex:
                if n is not None:
                    if r not in n:
                        found = False
                else:
                    found = False
            if found:
                findings.append((get_column_letter(i+1), n))
    return findings


def values_of(filename, columnname, nr=20):
    """
    #todo: Slim up
    """
    try:
        found = find_columns_by_regex(filename, columnname)
    except:
        print("Couldn't find columnname")
        raise
    else:
        ws = find_important_ws(filename)
        if len(found) != 1:
            print("Found more than one: {}\n\nTaking {}".format(found,
                                                                found[0][1]))
            letter = found[0][0]
            values = []
            import numpy as np
            for r in np.arange(1, 20):
                cell_name = letter+str(r)
                values.append(ws[cell_name].value)
            return values
        else:
            letter = found[0][0]
            values = []
            print("max rows", ws.max_row)
            import numpy as np
            for r in np.arange(1, 20):
                cell_name = letter+str(r)
                values.append(ws[cell_name].value)
            return values
